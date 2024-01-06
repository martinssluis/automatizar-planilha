import openpyxl

#criar uma planilha(book)
book = openpyxl.Workbook()
#como visualizar paginas existentes
print(book.sheetnames)
#como criar uima pagina
book.create_sheet('Computadores')
#como selecionar uma pagina
frutas_page = book['Computadores']
frutas_page.append(['Eletrônica','Memória Ram','Preço'])
frutas_page.append(['Computador 1','8gb Ram','R$2500'])
frutas_page.append(['Computador 2','16gb Ram','R$5500'])
frutas_page.append(['Computador 3','32gb Ram','R$8500'])
#salvar a planilha
book.save("Meus Computadores.xlsx")