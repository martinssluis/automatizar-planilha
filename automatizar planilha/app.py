import openpyxl

#criar uma planilha(book)
book = openpyxl.Workbook()
#como visualizar paginas existentes
print(book.sheetnames)
#como criar uima pagina
book.create_sheet('Frutas')
#como selecionar uma pagina
frutas_page = book['Frutas']
frutas_page.append(['Fruta','Quantidade','Preço'])
frutas_page.append(['Banana','5','R$3.90'])
frutas_page.append(['Fruta 2','2','R$15.90'])
frutas_page.append(['Fruta 3','10','R$30.90'])
frutas_page.append(['Fruta 4','2','R$50.50'])
#salvar a planilha
book.save("Planilha de Compras.xlsx")